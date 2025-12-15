from flask import current_app as app, jsonify, request, render_template, send_file
from .extensions import api, cache
import razorpay
import firebase_admin
from firebase_admin import auth as firebase_auth, credentials as firebase_credentials
import hmac
import hashlib
from flask_security import auth_required, roles_required, current_user,verify_password
from werkzeug.security import check_password_hash
import io
import openpyxl
import math
import requests
import time
import threading
from PIL import Image
import uuid
import sys
import os

from .models import db, User, Role, Restaurant ,RolesUsers,Order,OrderItem,MenuItem,Review,Category,RewardPoint,Coupon,TimeSlot
from .security import user_datastore
from .resources import RestaurantListAPI, RestaurantAPI, OrderAPI
from sqlalchemy import func,Date, or_
from datetime import datetime, date,timedelta

from sqlalchemy.orm import joinedload
import random
import string
import json

# --- ============================= ---
# --- CORE AUTHENTICATION API ROUTES ---
# --- ============================= ---

from flask import current_app as app, jsonify, request
from werkzeug.security import check_password_hash

# Make sure to import the user_datastore from your security file
from .security import user_datastore

# --- ✅ NEW: IMPORT FOR ONE-TIME DATABASE SETUP ---
from .create_initial_data import init_app as initialize_database
# --- END NEW IMPORT ---

# --- ✅ NEW: TEMPORARY ROUTE TO SET UP DATABASE ---
# --- Visit this URL ONCE after deploying to Render ---
@app.route('/api/admin/run-db-setup', methods=['GET'])
def temp_setup_database():
    """
    This is a temporary, one-time-use endpoint to initialize
    your roles and create the admin user on Render.
    """
    print("--- [TEMP SETUP] STARTING DATABASE INITIALIZATION ---")
    try:
        # Pass the actual app object to the function
        initialize_database(app._get_current_object()) 
        print("--- [TEMP SETUP] DATABASE INITIALIZATION COMPLETE ---")
        return jsonify({
            "status": "success",
            "message": "Database setup complete! Check Render logs for details. You can now log in with admin@crav.com / admin123"
        }), 200
    except Exception as e:
        error_message = f"Error during temp setup: {str(e)}"
        print(f"--- [TEMP SETUP] ERROR: {error_message} ---", file=sys.stderr)
        return jsonify({"status": "error", "message": error_message}), 500
# --- END NEW ROUTE ---

# Firebase Admin initialization is now handled centrally in `backend/firebase_admin_init.py`.
# The app factory (`createApp` in app.py) calls the initializer during startup.

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    if not data or not data.get('email') or not data.get('password'):
        return jsonify({"message": "Email and password are required"}), 400

    try:
        user = user_datastore.find_user(email=data.get('email'))

        # 👇 THIS IS THE CRITICAL CHANGE 👇
        # Use verify_password instead of check_password_hash
        if not user or not verify_password(data.get('password'), user.password):
            return jsonify({"message": "Invalid credentials"}), 401
            
        auth_token = user.get_auth_token()
        
        # User is authenticated
        return jsonify({
            "message": "Login Successful",
            "token": user.get_auth_token(),
            "user": {
                "id": user.id, "email": user.email, "name": user.name, "roles": [r.name for r in user.roles]
            }
        }), 200
        
    except Exception as e:
        # Rollback the session immediately if it's aborted (like with InFailedSqlTransaction)
        db.session.rollback()
        print(f"CRITICAL LOGIN ERROR: {e}")
        # Return a generic 500 error to the client
        return jsonify({"message": "An internal server error occurred. Please try again."}), 500


@app.route('/api/register', methods=['POST'])
def register_customer():
    data = request.get_json()
    email = data.get('email')
    if user_datastore.find_user(email=email):
        return jsonify({"message": "User already exists"}), 409
    
    # 👇 HASH THE PASSWORD ON REGISTRATION 👇
    pw = data.get('password')
    user = user_datastore.create_user(
        email=email,
        password=pw, # Flask-Security will hash
        name=data.get('name'),
        roles=['customer']
    )
    db.session.commit()

    # If Firebase Admin is initialized, also create a Firebase user so email/password works
    try:
        if firebase_admin._apps:
            try:
                fb_user = firebase_auth.create_user(email=email, password=pw)
                user.firebase_uid = fb_user.uid
                db.session.add(user)
                db.session.commit()
                app.logger.info('Created Firebase user for %s', email)
            except Exception as e:
                # Non-fatal: log but continue. Common reasons: email already exists in Firebase
                app.logger.warning('Failed to create Firebase user: %s', e)
    except Exception:
        # If firebase_admin not present or other issue, ignore
        pass

    return jsonify({"message": "Customer account created successfully"}), 201


@app.route('/api/auth/firebase', methods=['POST'])
def firebase_auth_exchange():
    """Accepts a Firebase ID token from the client, verifies it with
    Firebase Admin SDK, maps or creates a local user, and returns the
    application's auth token and user info.
    """
    data = request.get_json() or {}
    id_token = data.get('idToken')
    if not id_token:
        return jsonify({"message": "idToken is required"}), 400

    # Ensure firebase admin initialized
    if not firebase_admin._apps:
        app.logger.warning('Firebase admin not initialized when verifying token')
        return jsonify({"message": "Firebase not configured on server"}), 500

    try:
        decoded = firebase_auth.verify_id_token(id_token)
        email = decoded.get('email')
        name = decoded.get('name') or decoded.get('displayName') or email.split('@')[0]

        if not email:
            return jsonify({"message": "Firebase token does not contain an email"}), 400

        firebase_uid = decoded.get('uid') or decoded.get('user_id')
        photo = decoded.get('picture') or decoded.get('photoUrl')

        user = user_datastore.find_user(email=email)
        if not user:
            # Create a local user mapped to this Firebase identity
            random_pw = ''.join(random.choices(string.ascii_letters + string.digits, k=32))
            user = user_datastore.create_user(
                email=email,
                password=random_pw,
                name=name,
                roles=['customer']
            )
            # set firebase_uid and photo if available
            if firebase_uid:
                user.firebase_uid = firebase_uid
            if photo:
                user.photo_url = photo
            db.session.commit()
        else:
            # Update existing user record with Firebase details if missing or changed
            updated = False
            if firebase_uid and getattr(user, 'firebase_uid', None) != firebase_uid:
                user.firebase_uid = firebase_uid
                updated = True
            if photo and getattr(user, 'photo_url', None) != photo:
                user.photo_url = photo
                updated = True
            if updated:
                db.session.add(user)
                db.session.commit()

        # Return the same shape as the regular login endpoint
        return jsonify({
            "message": "Login Successful",
            "token": user.get_auth_token(),
            "user": {"id": user.id, "email": user.email, "name": user.name, "roles": [r.name for r in user.roles], "photo_url": getattr(user, 'photo_url', None)}
        }), 200

    except Exception as e:
        app.logger.exception('Firebase token verification failed')
        return jsonify({"message": "Invalid Firebase token"}), 401


@app.route('/api/config/firebase', methods=['GET'])
def get_firebase_config():
    """Return Firebase frontend config stored in server config (JSON string).
    This allows the frontend to initialize the Firebase SDK without bundling secrets.
    """
    cfg = app.config.get('FIREBASE_FRONTEND_CONFIG_JSON')
    if not cfg:
        return jsonify({}), 404
    try:
        parsed = json.loads(cfg)
        return jsonify(parsed), 200
    except Exception as e:
        app.logger.exception('Failed to parse FIREBASE_FRONTEND_CONFIG_JSON')
        return jsonify({}), 500

# --- ========================= ---
# --- RESTAURANT OWNER API ROUTES ---
# --- ========================= ---
#--- ✅ MODIFIED: RESTAURANT REGISTRATION ---
@app.route('/api/restaurant/register', methods=['POST'])
def register_restaurant():
    data = request.get_json()
    # ... (initial validation remains the same)
    
    # --- Add latitude and longitude to the new restaurant object ---
    try:
        owner = user_datastore.create_user(
            email=data.get('ownerEmail'),
            password=data.get('password'),
            name=data.get('ownerName'),
            roles=['owner']
        )
        db.session.commit()

        new_restaurant = Restaurant(
            owner_id=owner.id,
            name=data.get('restaurantName'),
            address=data.get('address'),
            city=data.get('city'),
            latitude=data.get('latitude'),   # <-- ADDED
            longitude=data.get('longitude'), # <-- ADDED
            is_verified=False
        )
        db.session.add(new_restaurant)
        db.session.commit()
        return jsonify({"message": "Restaurant submitted for verification!"}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"message": "An internal error occurred."}), 500
    

@app.route('/api/admin/restaurants', methods=['GET'])
@auth_required('token')
@roles_required('admin')
def get_all_restaurants():
    """ Fetches all restaurants for the admin panel, with optional filtering. """
    try:
        # Start with a base query
        query = db.session.query(Restaurant).options(joinedload(Restaurant.owner)).join(User, Restaurant.owner_id == User.id)

        # Get filter parameters from the request URL
        search_term = request.args.get('search', None)
        status_filter = request.args.get('status', None)

        # Apply search filter if provided
        if search_term:
            search_like = f"%{search_term}%"
            query = query.filter(
                or_(
                    Restaurant.name.ilike(search_like),
                    Restaurant.city.ilike(search_like),
                    User.email.ilike(search_like) # Assumes you're searching by owner email
                )
            )

        restaurants = query.all() # Execute the query
        
        # This part remains the same: determine status string based on flags
        restaurants_data = []
        for resto in restaurants:
            status = "Pending"
            if resto.is_verified and resto.owner and resto.owner.active:
                status = "Verified"
            elif resto.owner and not resto.owner.active:
                status = "Blocked"
            
            # This is server-side filtering for status, after getting all results
            # This is less efficient, but simpler than complex DB queries for status
            if status_filter and status_filter != 'All' and status != status_filter:
                continue

            restaurants_data.append({
                'id': resto.id,
                'name': resto.name,
                'ownerEmail': resto.owner.email if resto.owner else 'Owner Deleted',
                'city': resto.city,
                'status': status
            })
        
        return jsonify(restaurants_data), 200
    except Exception as e:
        print(f"Error fetching all restaurants: {e}")
        return jsonify({"message": "An error occurred on the server."}), 500

# (Removed duplicate Restaurant Dashboard, Fees, Orders, and Geolocation endpoints)


@app.route('/api/admin/dashboard', methods=['GET'])
@auth_required('token')
@roles_required('admin')
def admin_dashboard_stats():
    """ Gathers and returns all key metrics for the admin dashboard. """
    try:
        # Calculate total revenue from completed orders
        total_revenue = db.session.query(func.sum(Order.total_amount)).filter(Order.status == 'completed').scalar() or 0

        # Get total counts
        total_orders = db.session.query(func.count(Order.id)).scalar() or 0

        # --- MODIFIED QUERY FOR CUSTOMERS ---
        # This explicit join is more robust than the previous implicit one.
        total_customers = db.session.query(func.count(User.id)).join(RolesUsers, RolesUsers.user_id == User.id).join(Role, RolesUsers.role_id == Role.id).filter(Role.name == 'customer').scalar() or 0
        
        total_restaurants = db.session.query(func.count(Restaurant.id)).scalar() or 0

        # Eagerly load the 'owner' relationship to prevent extra queries and handle potential errors.
        pending_restaurants = Restaurant.query.options(joinedload(Restaurant.owner)).filter_by(is_verified=False).all()
        
        # Format the pending restaurants data, now with a safety check.
        pending_restaurants_data = [{
            'id': resto.id,
            'name': resto.name,
            # This check prevents a server crash if a restaurant has no owner.
            'ownerEmail': resto.owner.email if resto.owner else 'Owner Not Found',
            'city': resto.city
        } for resto in pending_restaurants]

        stats = {
            'totalRevenue': round(total_revenue, 2),
            'totalOrders': total_orders,
            'totalCustomers': total_customers,
            'totalRestaurants': total_restaurants
        }
        
        return jsonify({
            'stats': stats,
            'pendingRestaurants': pending_restaurants_data
        }), 200

    except Exception as e:
        print(f"Error fetching admin dashboard data: {e}")
        return jsonify({"message": "An error occurred on the server while fetching dashboard data."}), 500
# --- ================= ---
# --- ADMIN API ROUTES ---
# --- ================= ---



@app.route('/api/admin/restaurants/<int:id>/verify', methods=['PATCH'])
@auth_required('token')
@roles_required('admin')
def verify_restaurant(id):
    """ Approves a restaurant by setting its is_verified flag to True. """
    restaurant = Restaurant.query.get_or_404(id)
    restaurant.is_verified = True
    db.session.commit()
    return jsonify({"message": f"'{restaurant.name}' has been verified."}), 200

@app.route('/api/admin/restaurants/<int:id>/block', methods=['PATCH'])
@auth_required('token')
@roles_required('admin')
def block_restaurant(id):
    restaurant = Restaurant.query.get_or_404(id)
    if not restaurant.owner:
        return jsonify({"message": "Restaurant has no owner to block."}), 400
    restaurant.owner.active = False
    db.session.commit()
    return jsonify({"message": f"'{restaurant.name}' and its owner have been blocked."}), 200

@app.route('/api/admin/restaurants/<int:id>/unblock', methods=['PATCH'])
@auth_required('token')
@roles_required('admin')
def unblock_restaurant(id):
    restaurant = Restaurant.query.get_or_404(id)
    if not restaurant.owner:
        return jsonify({"message": "Restaurant has no owner to unblock."}), 400
    restaurant.owner.active = True
    db.session.commit()
    return jsonify({"message": f"'{restaurant.name}' and its owner have been unblocked."}), 200

@app.route('/api/admin/restaurants/<int:id>', methods=['DELETE'])
@auth_required('token')
@roles_required('admin')
def delete_restaurant(id):
    restaurant = Restaurant.query.get_or_404(id)
    db.session.delete(restaurant)
    db.session.commit()
    return jsonify({"message": f"'{restaurant.name}' has been permanently deleted."}), 200

# --- NEW: ADMIN ORDER MANAGEMENT ENDPOINTS ---

@app.route('/api/admin/orders', methods=['GET'])
@auth_required('token')
@roles_required('admin')
def admin_get_all_orders():
    """ Fetches all orders from all restaurants for the admin panel, with filtering. """
    try:
        query = Order.query.options(
            joinedload(Order.customer),
            joinedload(Order.restaurant)
        ).join(User, Order.user_id == User.id).join(Restaurant, Order.restaurant_id == Restaurant.id)

        search_term = request.args.get('search', None)
        status_filter = request.args.get('status', 'All')

        if search_term:
            search_like = f"%{search_term}%"
            query = query.filter(
                or_(
                    # Search by Order ID (requires casting to string)
                    Order.id.cast(db.String).ilike(search_like),
                    User.name.ilike(search_like),
                    Restaurant.name.ilike(search_like)
                )
            )
        
        if status_filter and status_filter != 'All':
            # Note: status is stored in lowercase in the db
            query = query.filter(Order.status == status_filter.lower())

        orders = query.order_by(Order.created_at.desc()).all()

        orders_data = [{
            'id': order.id,
            'customerName': order.customer.name if order.customer else 'N/A',
            'restaurantName': order.restaurant.name if order.restaurant else 'N/A',
            'date': order.created_at.strftime('%b %d, %Y'),
            'total': order.total_amount,
            'status': order.status.capitalize()
        } for order in orders]
        
        return jsonify(orders_data), 200
    except Exception as e:
        print(f"Error fetching all orders: {e}")
        return jsonify({"message": "An error occurred on the server."}), 500

@app.route('/api/admin/orders/<int:order_id>/refund', methods=['POST'])
@auth_required('token')
@roles_required('admin')
def admin_refund_order(order_id):
    """
    Processes a refund by updating the order's status to 'refunded'.
    In a real app, this would also interact with a payment gateway.
    """
    order = Order.query.get_or_404(order_id)
    
    # Prevent re-refunding an already refunded order
    if order.status == 'refunded':
        return jsonify({"message": f"Order #{order.id} has already been refunded."}), 400

    # Here you would add your payment gateway refund logic.
    # For this simulation, we'll assume it's successful.
    print(f"Admin initiated refund for Order #{order.id} amounting to ${order.total_amount}")
    
    # Update the order status
    order.status = 'refunded'
    db.session.commit()
    
    return jsonify({"message": f"Refund for Order #{order.id} has been successfully processed."}), 200
@app.route('/api/admin/reviews', methods=['GET'])
@auth_required('token')
@roles_required('admin')
def admin_get_all_reviews():
    """ Fetches all reviews from all restaurants for the admin panel. """
    reviews = Review.query.options(
        joinedload(Review.customer),
        joinedload(Review.restaurant)
    ).order_by(Review.created_at.desc()).all()

    reviews_data = [{
        'id': review.id,
        'customerName': review.customer.name if review.customer else 'N/A',
        'restaurantName': review.restaurant.name if review.restaurant else 'N/A',
        'rating': review.rating,
        'comment': review.comment,
        'date': review.created_at.strftime('%b %d, %Y')
    } for review in reviews]
    
    return jsonify(reviews_data), 200

@app.route('/api/admin/reviews/<int:review_id>', methods=['DELETE'])
@auth_required('token')
@roles_required('admin')
def admin_delete_review(review_id):
    """ Permanently deletes a review. """
    review = Review.query.get_or_404(review_id)
    db.session.delete(review)
    db.session.commit()
    return jsonify({"message": f"Review #{review.id} has been permanently deleted."}), 200

@app.route('/api/admin/coupons', methods=['GET', 'POST'])
@auth_required('token')
@roles_required('admin')
def admin_manage_platform_coupons():
    """ Fetches (GET) or creates (POST) platform-wide coupons. """
    if request.method == 'GET':
        # restaurant_id=None fetches platform-wide coupons
        coupons = Coupon.query.filter_by(restaurant_id=None).all()
        coupons_data = [{
            'id': c.id, 'code': c.code, 'type': c.discount_type,
            'value': c.discount_value, 'isActive': c.is_active
        } for c in coupons]
        return jsonify(coupons_data), 200

    if request.method == 'POST':
        data = request.get_json()
        new_coupon = Coupon(
            restaurant_id=None, # Explicitly set to None for platform-wide
            code=data['code'],
            discount_type=data['type'],
            discount_value=data['value'],
            is_active=data.get('isActive', True)
        )
        db.session.add(new_coupon)
        db.session.commit()
        return jsonify({"message": "Platform coupon created successfully."}), 201

@app.route('/api/admin/coupons/<int:coupon_id>', methods=['PUT', 'DELETE'])
@auth_required('token')
@roles_required('admin')
def admin_manage_specific_platform_coupon(coupon_id):
    """ Updates (PUT) or deletes (DELETE) a specific platform-wide coupon. """
    # Ensure we're only touching coupons with no restaurant_id
    coupon = Coupon.query.filter_by(id=coupon_id, restaurant_id=None).first_or_404()

    if request.method == 'PUT':
        data = request.get_json()
        coupon.code = data.get('code', coupon.code)
        coupon.discount_type = data.get('type', coupon.discount_type)
        coupon.discount_value = data.get('value', coupon.discount_value)
        coupon.is_active = data.get('isActive', coupon.is_active)
        db.session.commit()
        return jsonify({"message": "Platform coupon updated successfully."}), 200

    if request.method == 'DELETE':
        db.session.delete(coupon)
        db.session.commit()
        return jsonify({"message": "Platform coupon deleted successfully."}), 200


# In backend/routes.py

# ... (ensure all necessary imports are at the top of your file)
from sqlalchemy import func
from datetime import date, timedelta, datetime # <-- Make sure datetime is imported

@app.route('/api/admin/reports', methods=['GET'])
@auth_required('token')
@roles_required('admin')
def get_admin_reports():
    """ Gathers and returns platform-wide analytics for the admin reports page. """
    
    # --- Daily Revenue (Last 7 Days) ---
    seven_days_ago = datetime.now() - timedelta(days=7)
    
    # 1. Fetch all completed orders from the last 7 days.
    # We filter using a full datetime object, which is very reliable.
    recent_orders = db.session.query(Order.created_at, Order.total_amount).filter(
        Order.status == 'completed',
        Order.created_at >= seven_days_ago
    ).all()

    # 2. Use a Python dictionary to aggregate revenue by day.
    revenue_by_date = {}
    for i in range(7):
        current_date = date.today() - timedelta(days=i)
        revenue_by_date[current_date] = 0.0

    for order_time, revenue in recent_orders:
        order_date = order_time.date() # Extract just the date part from the datetime object
        if order_date in revenue_by_date:
            revenue_by_date[order_date] += revenue

    # 3. Format the data for the frontend.
    daily_revenue_data = []
    for d, r in sorted(revenue_by_date.items()):
        daily_revenue_data.append({
            'day': d.strftime('%b %d'),
            'revenue': round(float(r), 2)
        })

    # --- Top Performing Restaurants (This part is correct) ---
    top_restaurants_query = db.session.query(
        Restaurant.name,
        func.sum(Order.total_amount).label('total_revenue')
    ).join(Order, Restaurant.id == Order.restaurant_id)\
    .filter(Order.status == 'completed')\
    .group_by(Restaurant.name)\
    .order_by(func.sum(Order.total_amount).desc()).limit(5).all()

    top_restaurants_data = [{
        'rank': index + 1,
        'name': name,
        'revenue': round(float(total_revenue), 2)
    } for index, (name, total_revenue) in enumerate(top_restaurants_query)]

    return jsonify({
        'dailyRevenue': daily_revenue_data,
        'topRestaurants': top_restaurants_data
    }), 200


# --- NEW: USER MANAGEMENT ENDPOINTS ---

@app.route('/api/admin/users', methods=['GET'])
@auth_required('token')
@roles_required('admin')
def get_all_users():
    """ Fetches all customers with their order stats for the admin panel. """
    try:
        # Subquery to calculate total orders per user
        orders_subquery = db.session.query(
            Order.user_id,
            func.count(Order.id).label('total_orders'),
            func.sum(Order.total_amount).label('total_spent')
        ).group_by(Order.user_id).subquery()

        # Query all users with the 'customer' role
        customer_role = Role.query.filter_by(name='customer').first()
        query = db.session.query(
            User,
            func.coalesce(orders_subquery.c.total_orders, 0).label('total_orders'),
            func.coalesce(orders_subquery.c.total_spent, 0).label('total_spent')
        ).outerjoin(orders_subquery, User.id == orders_subquery.c.user_id).filter(User.roles.contains(customer_role))

        # Get search parameter from the request URL
        search_term = request.args.get('search', None)
        if search_term:
            search_like = f"%{search_term}%"
            query = query.filter(
                or_(
                    User.name.ilike(search_like),
                    User.email.ilike(search_like)
                )
            )

        users = query.all()

        users_data = [{
            'id': user.id,
            'name': user.name,
            'email': user.email,
            'totalOrders': total_orders,
            'totalSpent': round(float(total_spent), 2),
            'isBlocked': not user.active
        } for user, total_orders, total_spent in users]

        return jsonify(users_data), 200
    except Exception as e:
        print(f"Error fetching users: {e}")
        return jsonify({"message": "An error occurred on the server."}), 500

@app.route('/api/admin/users/<int:id>/block', methods=['PATCH'])
@auth_required('token')
@roles_required('admin')
def block_user(id):
    """ Blocks a user by setting their 'active' flag to False. """
    user = User.query.get_or_404(id)
    user.active = False
    db.session.commit()
    return jsonify({"message": f"User '{user.name}' has been blocked."}), 200

@app.route('/api/admin/users/<int:id>/unblock', methods=['PATCH'])
@auth_required('token')
@roles_required('admin')
def unblock_user(id):
    """ Unblocks a user by setting their 'active' flag to True. """
    user = User.query.get_or_404(id)
    user.active = True
    db.session.commit()
    return jsonify({"message": f"User '{user.name}' has been unblocked."}), 200

# --- ======================== ---
# --- CUSTOMER-SPECIFIC ROUTES ---
# --- ======================== ---

@app.route('/api/profile', methods=['GET', 'PUT'])
@auth_required('token')
@roles_required('customer') # Ensures only customers can access this
def manage_customer_profile():
    """
    Handles fetching and updating the logged-in customer's profile.
    """
    # The 'current_user' is provided by Flask-Security from the token
    user = current_user

    if request.method == 'GET':
        return jsonify({
            'id': user.id,
            'name': user.name,
            'email': user.email
        }), 200

    if request.method == 'PUT':
        data = request.get_json()
        if not data or not data.get('name'):
            return jsonify({"message": "Name is a required field."}), 400
        
        # Update the user's name
        user.name = data.get('name')
        db.session.commit()
        
        # Return the updated user info, so the frontend can update its state
        return jsonify({
            'message': 'Profile updated successfully!',
            'user': {
                'id': user.id,
                'name': user.name,
                'email': user.email,
                'roles': [r.name for r in user.roles]
            }
        }), 200



    if request.method == 'GET':
        # Your GET logic is fine and can remain as is.
        orders = Order.query.options(joinedload(Order.restaurant)).filter_by(user_id=current_user.id).order_by(Order.created_at.desc()).all()
        return jsonify([{'id': o.id, 'date': o.created_at.strftime('%b %d, %Y'), 'total': o.total_amount, 'status': o.status.capitalize(), 'restaurantName': o.restaurant.name if o.restaurant else 'N/A'} for o in orders]), 200

@app.route('/api/orders', methods=['POST'])
@auth_required('token')
@roles_required('customer')
def place_order():
    """
    Creates a new order, performing all necessary server-side validation.
    """
    data = request.get_json()
    restaurant_id = data.get('restaurant_id')
    if not restaurant_id:
        return jsonify({"message": "Restaurant ID is missing."}), 400

    restaurant = Restaurant.query.get_or_404(restaurant_id)

    # --- 1. Secure Server-Side Price Calculation ---
    subtotal = 0
    order_items_to_create = []
    cart_items = data.get('items', [])
    if not cart_items:
        return jsonify({"message": "Cannot place an empty order."}), 400

    for item_data in cart_items:
        menu_item = MenuItem.query.get(item_data.get('menu_item_id'))
        if not menu_item or not menu_item.is_available or menu_item.restaurant_id != restaurant.id:
            return jsonify({'message': "An item in your cart is invalid or unavailable."}), 400
        
        quantity = item_data.get('quantity')
        subtotal += menu_item.price * quantity
        order_items_to_create.append(OrderItem(
            menu_item_id=menu_item.id,
            quantity=quantity,
            price_at_order=menu_item.price
        ))

    # --- 2. Determine Order Type ---
    raw_order_type = (data.get('order_type') or 'takeaway').lower()
    if raw_order_type in ['dinein', 'dine_in', 'dining']:
        order_type = 'dine_in'
    elif raw_order_type in ['pickup', 'collect', 'takeaway']:
        order_type = 'takeaway'
    elif raw_order_type == 'delivery':
        order_type = 'delivery'
    else:
        order_type = 'takeaway'

    if order_type == 'delivery' and not data.get('delivery_address'):
        return jsonify({'message': 'Delivery address is required for delivery orders.'}), 400

    # --- 3. Secure Service Fee & Platform Fee ---
    platform_fee = restaurant.platform_fee if restaurant.platform_fee is not None else 7.0
    
    service_fee = 0.0
    if order_type == 'delivery':
        service_fee = restaurant.delivery_fee if restaurant.delivery_fee is not None else 50.0
    elif order_type == 'takeaway':
        service_fee = restaurant.takeaway_fee if restaurant.takeaway_fee is not None else 20.0
    elif order_type == 'dine_in':
        service_fee = restaurant.dine_in_fee if restaurant.dine_in_fee is not None else 10.0

    # Discount Logic
    discount_amount = 0
    coupon_id = None
    coupon_code = data.get('coupon_code')
    
    if coupon_code:
        coupon = Coupon.query.filter(
            Coupon.code == coupon_code,
            Coupon.is_active == True,
            or_(Coupon.restaurant_id == None, Coupon.restaurant_id == restaurant.id)
        ).first()
        if coupon:
            coupon_id = coupon.id
            if coupon.discount_type == 'Percentage':
                discount_amount = (subtotal * coupon.discount_value) / 100
            else: 
                discount_amount = coupon.discount_value
    
    # Final Calculation
    gross_total = subtotal + platform_fee + service_fee
    final_total = max(0, gross_total - discount_amount)

    # --- 4. Handle Scheduled Time ---
    scheduled_time_obj = None 
    if data.get('scheduled_time'):
        try:
            iso_string = data['scheduled_time']
            if iso_string.endswith('Z'):
                iso_string = iso_string[:-1] + "+00:00"
            scheduled_time_obj = datetime.fromisoformat(iso_string)
        except (ValueError, TypeError):
            return jsonify({"message": "Invalid format for scheduled time."}), 400

    # --- 5. Create the Order Object ---
    otp = ''.join(random.choices(string.digits, k=6))
    qr_payload = ''.join(random.choices(string.ascii_letters + string.digits, k=20))
    table_number = data.get('table_number') if order_type == 'dine_in' else None

    # Delivery Address
    delivery_address = data.get('delivery_address') if order_type == 'delivery' else None

    new_order = Order(
        user_id=current_user.id,
        restaurant_id=restaurant.id,
        total_amount=round(final_total, 2),
        order_type=order_type,
        table_number=table_number,
        status='placed',
        otp=otp,
        qr_payload=qr_payload,
        items=order_items_to_create,
        coupon_id=coupon_id,
        discount_amount=round(discount_amount, 2),
        is_scheduled=bool(scheduled_time_obj),
        scheduled_time=scheduled_time_obj,
        delivery_address=delivery_address
    )

    db.session.add(new_order)
    db.session.commit()
    try:
        cache.delete_memoized(get_restaurant_details, restaurant.id)
        cache.delete_memoized(get_featured_restaurants)
    except Exception:
        pass

    return jsonify({'message': 'Order placed successfully!', 'order_id': new_order.id}), 201


# -------------------- Razorpay Payment Endpoints --------------------
@app.route('/api/payments/create', methods=['POST'])
@auth_required('token')
@roles_required('customer')
def create_razorpay_order():
    """Create a Razorpay order for a given internal Order ID and return the razorpay_order_id and key for checkout."""
    data = request.get_json() or {}
    order_id = data.get('order_id')
    if not order_id:
        return jsonify({'message': 'order_id is required'}), 400

    order = Order.query.get_or_404(order_id)
    # Only allow creating payment for the user who owns the order
    if order.user_id != current_user.id:
        return jsonify({'message': 'Unauthorized to create payment for this order'}), 403

    # Get Razorpay credentials
    key_id = app.config.get('RAZORPAY_KEY_ID')
    key_secret = app.config.get('RAZORPAY_KEY_SECRET')
    
    amount_paisa = int(round(order.total_amount * 100))
    
    # Try with Razorpay if credentials exist, otherwise use mock (for development)
    if key_id and key_secret:
        try:
            client = razorpay.Client(auth=(key_id, key_secret))
            rp_order = client.order.create({
                'amount': amount_paisa,
                'currency': 'INR',
                'receipt': str(order.id),
                'payment_capture': 1
            })
            razorpay_order_id = rp_order.get('id')
        except Exception as e:
            print(f"Razorpay error: {e}. Using mock order for development.")
            # Fallback to mock for development
            razorpay_order_id = f'order_{order.id}_{int(time.time())}'
    else:
        # Development mode: create a mock order
        razorpay_order_id = f'order_{order.id}_{int(time.time())}'
    
    # Save razorpay_order_id to our order
    order.razorpay_order_id = razorpay_order_id
    order.payment_amount = order.total_amount
    db.session.commit()

    return jsonify({
        'razorpay_order_id': razorpay_order_id,
        'razorpay_key': key_id or 'test-key',
        'amount': amount_paisa
    }), 200


@app.route('/api/payments/verify', methods=['POST'])
@auth_required('token')
@roles_required('customer')
def verify_razorpay_payment():
    """Verify the payment signature returned from Razorpay checkout and mark order paid."""
    data = request.get_json() or {}
    razorpay_order_id = data.get('razorpay_order_id')
    razorpay_payment_id = data.get('razorpay_payment_id')
    razorpay_signature = data.get('razorpay_signature')
    order_id = data.get('order_id')

    if not all([razorpay_order_id, razorpay_payment_id, razorpay_signature, order_id]):
        return jsonify({'message': 'Missing payment verification fields'}), 400

    order = Order.query.get_or_404(order_id)
    if order.user_id != current_user.id:
        return jsonify({'message': 'Unauthorized'}), 403

    # For development mode with mock orders, skip signature verification
    key_secret = app.config.get('RAZORPAY_KEY_SECRET')
    if razorpay_order_id.startswith('order_'):
        # This is a mock order, accept it as-is
        print(f"Development mode: Accepting mock payment for order {order_id}")
    elif key_secret:
        # Verify signature: HMAC_SHA256(order_id|payment_id, key_secret)
        payload = f"{razorpay_order_id}|{razorpay_payment_id}".encode()
        secret = key_secret.encode()
        generated_sig = hmac.new(secret, payload, hashlib.sha256).hexdigest()

        if not hmac.compare_digest(generated_sig, razorpay_signature):
            return jsonify({'message': 'Invalid payment signature'}), 400

    # Mark payment success
    order.razorpay_payment_id = razorpay_payment_id
    order.payment_status = 'paid'
    
    # ✅ FIX: Status should be 'placed' so it appears in the Restaurant Dashboard
    # 'completed' is ONLY for when the order is picked up/delivered.
    order.status = 'placed' 
    order.acceptance_status = 'pending'
    
    db.session.commit()

    try:
        cache.delete_memoized(get_restaurant_details, order.restaurant_id)
        cache.delete_memoized(get_featured_restaurants)
    except Exception:
        pass

    return jsonify({'message': 'Payment verified and order completed.'}), 200
@app.route('/api/payments/webhook', methods=['POST'])
def razorpay_webhook():
    """Handle Razorpay webhooks. Verify signature using RAZORPAY_KEY_SECRET."""
    payload = request.get_data()
    signature = request.headers.get('X-Razorpay-Signature')
    secret = app.config.get('RAZORPAY_KEY_SECRET') or ''

    if not signature:
        return jsonify({'message': 'Missing signature'}), 400

    computed = hmac.new(secret.encode(), payload, hashlib.sha256).hexdigest()
    if not hmac.compare_digest(computed, signature):
        return jsonify({'message': 'Invalid signature'}), 400

    event = request.get_json()
    # Example: handle payment.captured
    try:
        event_type = event.get('event')
        if event_type == 'payment.captured':
            payment = event.get('payload', {}).get('payment', {}).get('entity', {})
            razorpay_payment_id = payment.get('id')
            razorpay_order_id = payment.get('order_id')
            # Find our order by receipt or razorpay_order_id
            order = None
            if razorpay_order_id:
                order = Order.query.filter_by(razorpay_order_id=razorpay_order_id).first()
            if not order and payment.get('notes'):
                # fallback if you stored receipt in notes
                pass

            if order:
                order.razorpay_payment_id = razorpay_payment_id
                order.payment_status = 'paid'
                
                # ✅ FIX: Webhook also sets status to 'placed' for visibility
                order.status = 'placed'
                order.acceptance_status = 'pending'
                
                db.session.commit()

    except Exception as e:
        print(f"Error processing webhook: {e}")

    return jsonify({'message': 'Webhook processed'}), 200

# -------------------- End Razorpay Endpoints --------------------


# In backend/routes.py

@app.route('/api/coupons/applicable/<int:restaurant_id>', methods=['GET'])
@auth_required('token')
@roles_required('customer')
def get_applicable_coupons(restaurant_id):
    """
    Fetches the top 3 active coupons for a specific restaurant, 
    including platform-wide ones. Returns full details.
    """
    try:
        # ✅ THE FIX: Added order_by and limit, and querying the full object
        coupons = Coupon.query.filter(
            Coupon.is_active == True,
            or_(Coupon.restaurant_id == None, Coupon.restaurant_id == restaurant_id)
        ).order_by(Coupon.id.desc()).limit(3).all()
        
        # ✅ THE FIX: Return more details for a better frontend display
        coupons_data = [{
            'code': c.code,
            'discount_type': c.discount_type,
            'discount_value': c.discount_value
        } for c in coupons]
        
        return jsonify(coupons_data), 200
        
    except Exception as e:
        print(f"Error fetching applicable coupons: {e}")
        return jsonify({"message": "Could not retrieve coupons."}), 500
@app.route('/api/coupons/apply', methods=['POST'])
@auth_required('token')
@roles_required('customer')
def apply_coupon():
    """
    Validates a coupon code and calculates the discount amount.
    """
    data = request.get_json()
    code = data.get('code')
    subtotal = data.get('subtotal')
    restaurant_id = data.get('restaurant_id')

    if not all([code, subtotal, restaurant_id]):
        return jsonify({"message": "Missing coupon code, subtotal, or restaurant ID."}), 400

    try:
        # Find a coupon that matches the code, is active, AND is either platform-wide or for the correct restaurant
        coupon = Coupon.query.filter(
            Coupon.code == code,
            Coupon.is_active == True,
            or_(Coupon.restaurant_id == None, Coupon.restaurant_id == restaurant_id)
        ).first()

        if not coupon:
            return jsonify({"message": "Invalid or expired coupon code."}), 404

        # Calculate the discount
        discount_amount = 0
        if coupon.discount_type == 'Percentage':
            discount_amount = (subtotal * coupon.discount_value) / 100
        elif coupon.discount_type == 'Fixed':
            discount_amount = coupon.discount_value
        
        # Ensure discount doesn't exceed the subtotal
        discount_amount = min(discount_amount, subtotal)

        return jsonify({
            "message": "Coupon applied successfully!",
            "discount": round(discount_amount, 2)
        }), 200

    except Exception as e:
        print(f"Error applying coupon: {e}")
        return jsonify({"message": "An error occurred while applying the coupon."}), 500

# --- UPDATED FAVORITES ENDPOINT ---
@app.route('/api/favorites', methods=['GET'])
@auth_required('token')
@roles_required('customer')
def get_favorites():
    """ Fetches all favorite restaurants for the logged-in customer. """
    favorites_data = []
    for resto in current_user.favorites:
        avg_rating = db.session.query(func.avg(Review.rating)).filter(Review.restaurant_id == resto.id).scalar() or 0.0
        review_count = db.session.query(func.count(Review.id)).filter(Review.restaurant_id == resto.id).scalar() or 0
        favorites_data.append({
            'id': resto.id, 'name': resto.name, 'cuisine': 'Local Cuisine',
            'rating': round(float(avg_rating), 1), 'reviews': review_count,
            'image': f'https://placehold.co/600x400/E65100/FFF?text={resto.name.replace(" ", "+")}'
        })
    return jsonify(favorites_data), 200

@app.route('/api/favorites/<int:restaurant_id>', methods=['POST', 'DELETE'])
@auth_required('token')
@roles_required('customer')
def manage_favorite(restaurant_id):
    restaurant = Restaurant.query.get_or_404(restaurant_id)
    if request.method == 'POST':
        if restaurant not in current_user.favorites:
            current_user.favorites.append(restaurant)
            db.session.commit()
        return jsonify({"message": "Favorite added."}), 201
    if request.method == 'DELETE':
        if restaurant in current_user.favorites:
            current_user.favorites.remove(restaurant)
            db.session.commit()
        return jsonify({"message": "Favorite removed."}), 200

# --- NEW: RESTAURANT LISTING & DETAIL ENDPOINTS ---

@app.route('/api/restaurants/featured', methods=['GET'])
@cache.cached(timeout=300)
def get_featured_restaurants():
    try:
        restaurants = Restaurant.query.filter_by(is_verified=True, is_active=True).limit(6).all()
        restaurants_data = []
        for resto in restaurants:
            # Calculate stats
            stats = db.session.query(
                func.avg(Review.rating).label('avg_rating'),
                func.count(Review.id).label('review_count')
            ).filter(Review.restaurant_id == resto.id).first()
            
            # ✅ START: LOGIC TO USE GALLERY IMAGE OR FALLBACK
            image_url = f'https://placehold.co/600x400/E65100/FFF?text={resto.name.replace(" ", "+")}'
            if resto.gallery and len(resto.gallery) > 0:
                image_url = resto.gallery[0]
            # ✅ END: LOGIC TO USE GALLERY IMAGE

            restaurants_data.append({
                'id': resto.id, 'name': resto.name, 'cuisine': 'Local Favorites',
                'rating': round(float(stats.avg_rating or 0), 1),
                'reviews': stats.review_count or 0,
                'image': image_url # Use the dynamic image_url
            })
        
        return jsonify(restaurants_data), 200

    except Exception as e:
        # CRITICAL FIX: Rollback the transaction to unblock the database
        db.session.rollback()
        print(f"CRITICAL FEATURED RESTAURANTS ERROR: {e}")
        # Return an empty list on error to prevent a 500 crash page
        return jsonify([]), 500



@app.route('/api/restaurants/<int:restaurant_id>', methods=['GET'])
# Removed cache to ensure fee updates reflect immediately
def get_restaurant_details(restaurant_id):
    restaurant = Restaurant.query.options(joinedload(Restaurant.categories).joinedload(Category.menu_items)).get_or_404(restaurant_id)
    
    stats = db.session.query(
        func.avg(Review.rating).label('avg_rating'),
        func.count(Review.id).label('review_count')
    ).filter(Review.restaurant_id == restaurant.id).first()

    categories_data = [{'id': cat.id, 'name': cat.name, 'menu_items': [{'id': item.id, 'name': item.name, 'description': item.description, 'price': item.price, 'is_available': item.is_available, 'image': item.image_url or f'https://placehold.co/600x400/E65100/FFF?text={item.name.replace(" ", "+")}' } for item in cat.menu_items]} for cat in restaurant.categories]
    
    restaurant_data = {
        'id': restaurant.id, 'name': restaurant.name, 'description': restaurant.description, 'address': restaurant.address, 'city': restaurant.city, 'cuisine': 'Local Favorites', 
        'rating': round(float(stats.avg_rating or 0), 1),
        'reviews': stats.review_count or 0,
        'categories': categories_data,
        
        # --- ✅ FIX: Include Fees ---
        'delivery_fee': restaurant.delivery_fee,
        'takeaway_fee': restaurant.takeaway_fee,
        'dine_in_fee': restaurant.dine_in_fee,
        'platform_fee': restaurant.platform_fee
    }
    
    # ✅ FIX: Add cache-control headers to prevent caching in production
    response = jsonify(restaurant_data)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response, 200

# In backend/routes.py

# In backend/routes.py

@app.route('/api/orders', methods=['GET'])
@auth_required('token')
@roles_required('customer')
def get_order_history():
    """ Fetches the order history for the logged-in customer. """
    try:
        # We use joinedload for both restaurant and the new review relationship for efficiency
        orders = Order.query.options(
                joinedload(Order.restaurant),
                joinedload(Order.review) # <-- Eagerly load the review
            )\
            .filter_by(user_id=current_user.id)\
            .order_by(Order.created_at.desc())\
            .all()

        orders_data = [{
            'id': o.id,
            'date': o.created_at.strftime('%b %d, %Y'),
            'total': o.total_amount,
            'status': o.status.capitalize(),
            'restaurantName': o.restaurant.name if o.restaurant else 'N/A',
            # ✅ FIX: Include payment/acceptance info
            'transaction_id': o.razorpay_payment_id,
            'acceptance_status': o.acceptance_status or 'pending',
            # ✅ THE FIX: Check if the review object exists for this order
            'has_review': bool(o.review) 
        } for o in orders]
        
        return jsonify(orders_data), 200
    except Exception as e:
        print(f"Error fetching order history: {e}")
        return jsonify({"message": "An error occurred while fetching your orders."}), 500


        
# --- REWARDS ENDPOINT ---
@app.route('/api/rewards', methods=['GET'])
@auth_required('token')
@roles_required('customer')
def get_rewards_data():
    total_points = db.session.query(func.sum(RewardPoint.points)).filter_by(user_id=current_user.id).scalar() or 0
    history = RewardPoint.query.filter_by(user_id=current_user.id).order_by(RewardPoint.created_at.desc()).all()
    history_data = [{'id': item.id, 'reason': item.reason, 'points': item.points, 'date': item.created_at.strftime('%b %d, %Y'), 'type': item.transaction_type} for item in history]
    return jsonify({'points_balance': total_points, 'history': history_data}), 200

@app.route('/api/restaurant/dashboard', methods=['GET'])
@auth_required('token')
@roles_required('owner')
def restaurant_dashboard_stats():
    """ Gathers and returns all key metrics for the restaurant owner's dashboard. """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first()
    
    if not restaurant:
        return jsonify({"message": "No restaurant profile found for this account. Please contact support if you believe this is an error."}), 404
    
    try:
        today = date.today()

        # --- Calculate Stats ---
        # Today's Revenue
        todays_revenue = db.session.query(func.sum(Order.total_amount))\
            .filter(Order.restaurant_id == restaurant.id, func.cast(Order.created_at, Date) == today).scalar() or 0.0

        # Today's Orders
        todays_orders = db.session.query(func.count(Order.id))\
            .filter(Order.restaurant_id == restaurant.id, func.cast(Order.created_at, Date) == today).scalar() or 0
            
        # Pending Orders (Placed or Preparing)
        pending_orders = db.session.query(func.count(Order.id))\
            .filter(Order.restaurant_id == restaurant.id, Order.status.in_(['placed', 'preparing'])).scalar() or 0

        # Recent Orders
        recent_orders_query = Order.query.filter_by(restaurant_id=restaurant.id)\
            .order_by(Order.created_at.desc()).limit(5).all()
        
        recent_orders_data = [{
            'id': order.id,
            'customerName': order.customer.name if order.customer else 'Unknown',
            'items': len(order.items),
            'total': order.total_amount,
            'status': order.status.capitalize()
        } for order in recent_orders_query]

        # Most Popular Items
        popular_items_query = db.session.query(
                MenuItem.name,
                func.count(OrderItem.id).label('order_count')
            ).join(OrderItem, MenuItem.id == OrderItem.menu_item_id)\
            .filter(MenuItem.restaurant_id == restaurant.id)\
            .group_by(MenuItem.name)\
            .order_by(func.count(OrderItem.id).desc()).limit(5).all()

        popular_items_data = [{'name': name, 'orders': count} for name, count in popular_items_query]

        stats = {
            'todaysRevenue': round(todays_revenue, 2),
            'todaysOrders': todays_orders,
            'pendingOrders': pending_orders,
        }

        return jsonify({
            'stats': stats,
            'recentOrders': recent_orders_data,
            'popularItems': popular_items_data
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"Error in dashboard stats: {e}")
        return jsonify({"message": "Failed to load dashboard data."}), 500

@app.route('/api/restaurant/fees', methods=['GET', 'PUT'])
@auth_required('token')
@roles_required('owner')
def manage_restaurant_fees():
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    
    if request.method == 'GET':
        return jsonify({
            'delivery_fee': restaurant.delivery_fee,
            'takeaway_fee': restaurant.takeaway_fee,
            'dine_in_fee': restaurant.dine_in_fee,
            'platform_fee': restaurant.platform_fee
        }), 200
        
    if request.method == 'PUT':
        data = request.get_json()
        if 'delivery_fee' in data: restaurant.delivery_fee = float(data['delivery_fee'])
        if 'takeaway_fee' in data: restaurant.takeaway_fee = float(data['takeaway_fee'])
        if 'dine_in_fee' in data: restaurant.dine_in_fee = float(data['dine_in_fee'])
        if 'platform_fee' in data: restaurant.platform_fee = float(data['platform_fee'])
        
        db.session.commit()
        return jsonify({"message": "Fee settings updated successfully."}), 200

# --- NEW: ORDER QUEUE ENDPOINTS ---

@app.route('/api/restaurant/orders', methods=['GET'])
@auth_required('token')
@roles_required('owner')
def get_restaurant_orders():
    """ 
    Fetches all active orders for the owner's restaurant.
    Includes scheduled time for relevant orders.
    """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    
    # The query correctly fetches all non-finalized orders
    orders = Order.query.filter(
        Order.restaurant_id == restaurant.id,
        Order.status.notin_(['completed', 'cancelled', 'rejected', 'refunded'])
    ).order_by(Order.created_at.asc()).all()

    orders_data = []
    for order in orders:
        items_data = [{'name': item.menu_item.name if item.menu_item else 'Unknown Item', 'quantity': item.quantity} for item in order.items]
        
        # Convert UTC creation time to IST for display
        ist_created_time = order.created_at + timedelta(hours=5, minutes=30)

        order_info = {
            'id': order.id,
            'customerName': order.customer.name,
            'createdAt': ist_created_time.strftime('%I:%M %p'),
            'status': order.status,
            'order_type': order.order_type,
            'items': items_data,
            'is_scheduled': order.is_scheduled, # Pass the flag
            'scheduled_date': None,
            'scheduled_time': None
        }

        # If the order is scheduled, add the formatted IST time
        if order.is_scheduled and order.scheduled_time:
            ist_scheduled_time = order.scheduled_time + timedelta(hours=5, minutes=30)
            order_info['scheduled_date'] = ist_scheduled_time.strftime('%b %d, %Y')
            order_info['scheduled_time'] = ist_scheduled_time.strftime('%I:%M %p')

        orders_data.append(order_info)
        
    return jsonify(orders_data), 200



@app.route('/api/restaurant/orders/<int:order_id>/status', methods=['PATCH'])
@auth_required('token')
@roles_required('owner')
def update_order_status(order_id):
    """ Updates the status of an order (e.g., accept, reject, prepare, ready). """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    order = Order.query.get_or_404(order_id)
    
    if order.restaurant_id != restaurant.id:
        return jsonify({"message": "Unauthorized to modify this order."}), 403

    data = request.get_json()
    new_status = data.get('status')
    
    # 'completed' is no longer allowed here; it must go through verification.
    allowed_statuses = ['preparing', 'ready', 'rejected']
    if new_status not in allowed_statuses:
        return jsonify({"message": f"Invalid status '{new_status}'."}), 400
        
    order.status = new_status
    
    # ✅ FIX: Update acceptance_status based on the workflow status
    if new_status == 'preparing':
        order.acceptance_status = 'accepted'
    elif new_status == 'rejected':
        order.acceptance_status = 'rejected'
        
    db.session.commit()
    try:
        cache.delete_memoized(get_restaurant_details, restaurant.id)
        cache.delete_memoized(get_featured_restaurants)
    except Exception:
        pass
    
    return jsonify({"message": f"Order #{order.id} has been updated to '{new_status}'."}), 200


@app.route('/api/restaurant/orders/<int:order_id>/pickup', methods=['PATCH'])
@auth_required('token')
@roles_required('owner')
def set_pickup_ready(order_id):
    """Marks an order as ready for pickup (or unmarks it)."""
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    order = Order.query.get_or_404(order_id)
    if order.restaurant_id != restaurant.id:
        return jsonify({"message": "Unauthorized to modify this order."}), 403

    data = request.get_json() or {}
    ready = data.get('pickup_ready', True)
    order.pickup_ready = bool(ready)
    if ready:
        order.status = 'ready'
    db.session.commit()
    try:
        cache.delete_memoized(get_restaurant_details, restaurant.id)
        cache.delete_memoized(get_featured_restaurants)
    except Exception:
        pass
    return jsonify({"message": f"Order #{order.id} pickup_ready set to {order.pickup_ready}."}), 200

# ✅ START: NEW OTP VERIFICATION ROUTE
@app.route('/api/restaurant/orders/<int:order_id>/verify', methods=['POST'])
@auth_required('token')
@roles_required('owner')
def verify_order(order_id):
    """ Verifies an OTP and marks the order as 'completed'. """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    order = Order.query.get_or_404(order_id)

    # Security and status checks
    if order.restaurant_id != restaurant.id:
        return jsonify({"message": "Unauthorized."}), 403
    if order.status != 'ready':
        return jsonify({"message": "Order must be marked as 'Ready' before verification."}), 400

    data = request.get_json()
    otp_submitted = data.get('otp')
    if not otp_submitted:
        return jsonify({"message": "OTP is required."}), 400

    # The core logic: compare OTPs
    print(f"DEBUG: Comparing OTPs - Input: '{otp_submitted}' stored: '{order.otp}'")
    if str(otp_submitted).strip() == str(order.otp).strip():
        order.status = 'completed'
        db.session.commit()

        # ✅ START: This is the new logic to trigger the background task
        # We start a new thread to run our clearing function.
        # This does NOT block the server from sending the response immediately.
        thread = threading.Thread(
            target=clear_otp_after_delay,
            args=(app.app_context(), order.id)
        )
        thread.start()
        # ✅ END: New logic

        return jsonify({"message": f"Order #{order.id} verified and completed successfully!"}), 200
    else:
        return jsonify({"message": "Invalid OTP. Please try again."}), 400

def clear_otp_after_delay(app_context, order_id):
    """Waits for a delay and then clears the OTP for a given order."""
    with app_context:
        print(f"Starting background task to clear OTP for order {order_id} in 20 minutes.")
        time.sleep(1200) # Wait for 20 minutes (1200 seconds)
        try:
            order = Order.query.get(order_id)
            if order:
                order.otp = None # Clear the OTP field
                db.session.commit()
                print(f"Successfully cleared OTP for order {order_id}.")
        except Exception as e:
            print(f"Error in background task for order {order_id}: {e}")

# --- NEW: ORDER DETAILS ENDPOINT ---
@app.route('/api/orders/<int:order_id>', methods=['GET'])
@auth_required('token')
@roles_required('customer')
def get_order_details(order_id):
    """ Fetches details for a single order. """
    try:
        order = Order.query.options(
            joinedload(Order.restaurant),
            joinedload(Order.items).joinedload(OrderItem.menu_item)
        ).get_or_404(order_id)

        if order.user_id != current_user.id:
            return jsonify({"message": "Unauthorized access to this order."}), 403

        # Format items
        items_data = []
        for item in order.items:
            items_data.append({
                'id': item.id,
                'name': item.menu_item.name if item.menu_item else 'Unknown Item',
                'quantity': item.quantity,
                'price': item.price_at_order,
                'total': item.price_at_order * item.quantity
            })
            
        # Format the scheduled date/time for display
        scheduled_date_str = None
        scheduled_time_str = None
        if order.is_scheduled and order.scheduled_time:
             # Convert to IST (UTC+5:30) for display
            ist_time = order.scheduled_time + timedelta(hours=5, minutes=30)
            scheduled_date_str = ist_time.strftime('%b %d, %Y')
            scheduled_time_str = ist_time.strftime('%I:%M %p')

        order_data = {
            'id': order.id,
            'date': order.created_at.strftime('%b %d, %Y'),
            'total': order.total_amount,
            'status': order.status.capitalize(),
            'restaurantName': order.restaurant.name if order.restaurant else 'N/A',
            'order_type': order.order_type,
            'otp': order.otp,
            'qr_payload': order.qr_payload,
            'items': items_data,
            'is_scheduled': order.is_scheduled,
            'scheduled_date': scheduled_date_str,
            'scheduled_time': scheduled_time_str,
            # ✅ FIX: Include payment/acceptance info for details page
            'razorpay_payment_id': order.razorpay_payment_id,
            'acceptance_status': order.acceptance_status or 'pending'
        }

        return jsonify(order_data), 200
    except Exception as e:
        print(f"Error fetching order details: {e}")
        return jsonify({"message": "An error occurred."}), 500
# --- NEW: MENU MANAGEMENT ENDPOINTS ---

@app.route('/api/restaurant/menu', methods=['GET'])
@auth_required('token')
@roles_required('owner')
def get_restaurant_menu():
    """ Fetches all categories and menu items for the owner's restaurant. """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    categories = Category.query.options(joinedload(Category.menu_items)).filter_by(restaurant_id=restaurant.id).all()
    
    categories_data = [{
        'id': cat.id,
        'name': cat.name,
        'menu_items': [{
            'id': item.id,
            'name': item.name,
            'description': item.description,
            'price': item.price,
            'is_available': item.is_available,
            'image': item.image_url or f'https://placehold.co/100x100/E65100/FFF?text={item.name.replace(" ", "+")}'
        } for item in cat.menu_items]
    } for cat in categories]
    
    return jsonify(categories_data), 200

@app.route('/api/restaurant/menu-items', methods=['POST'])
@auth_required('token')
@roles_required('owner')
def create_menu_item():
    """ Creates a new menu item for the owner's restaurant. """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    data = request.get_json()
    category = Category.query.filter_by(id=data.get('category_id'), restaurant_id=restaurant.id).first()
    if not category:
        return jsonify({"message": "Invalid category."}), 400

    new_item = MenuItem(
        restaurant_id=restaurant.id,
        name=data.get('name'),
        price=data.get('price'),
        category_id=data.get('category_id'),
        description=data.get('description', ''),
        image_url=data.get('image', '')
    )
    db.session.add(new_item)
    db.session.commit()
    # Invalidate caches affected by menu changes
    try:
        cache.delete_memoized(get_restaurant_details, restaurant.id)
        cache.delete_memoized(get_featured_restaurants)
    except Exception:
        pass
    return jsonify({"message": "Menu item created successfully."}), 201

@app.route('/api/restaurant/menu-items/<int:item_id>', methods=['PUT', 'DELETE'])
@auth_required('token')
@roles_required('owner')
def manage_menu_item(item_id):
    """ Updates (PUT) or deletes (DELETE) a specific menu item. """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    item = MenuItem.query.get_or_404(item_id)
    if item.restaurant_id != restaurant.id:
        return jsonify({"message": "Unauthorized to modify this item."}), 403
        
    if request.method == 'PUT':
        data = request.get_json()
        item.name = data.get('name', item.name)
        item.description = data.get('description', item.description)
        item.price = data.get('price', item.price)
        item.image_url = data.get('image', item.image_url)
        item.category_id = data.get('category_id', item.category_id)
        db.session.commit()
        try:
            cache.delete_memoized(get_restaurant_details, restaurant.id)
            cache.delete_memoized(get_featured_restaurants)
        except Exception:
            pass
        return jsonify({"message": "Menu item updated successfully."}), 200

    if request.method == 'DELETE':
        db.session.delete(item)
        db.session.commit()
        try:
            cache.delete_memoized(get_restaurant_details, restaurant.id)
            cache.delete_memoized(get_featured_restaurants)
        except Exception:
            pass
        return jsonify({"message": "Menu item deleted successfully."}), 200

@app.route('/api/restaurant/menu-items/<int:item_id>/availability', methods=['PATCH'])
@auth_required('token')
@roles_required('owner')
def toggle_item_availability(item_id):
    """ Toggles the is_available status of a menu item. """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    item = MenuItem.query.get_or_404(item_id)
    if item.restaurant_id != restaurant.id:
        return jsonify({"message": "Unauthorized."}), 403
    
    data = request.get_json()
    if 'is_available' in data:
        item.is_available = data['is_available']
        db.session.commit()
        try:
            cache.delete_memoized(get_restaurant_details, restaurant.id)
            cache.delete_memoized(get_featured_restaurants)
        except Exception:
            pass
    
    return jsonify({"message": f"'{item.name}' availability updated."}), 200

@app.route('/api/restaurant/profile', methods=['GET', 'PUT'])
@auth_required('token')
@roles_required('owner')
def manage_restaurant_profile():
    """ Fetches (GET) or updates (PUT) the profile for the owner's restaurant. """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()

    if request.method == 'GET':
        profile_data = {
            'name': restaurant.name,
            'description': restaurant.description,
            'address': restaurant.address,
            'city': restaurant.city,
            'isActive': restaurant.is_active,
            'openingHours': '9:00 AM - 10:00 PM', # Placeholder
            'gallery': restaurant.gallery or [] # Fetch gallery from DB
        }
        return jsonify(profile_data), 200

    if request.method == 'PUT':
        data = request.get_json()
        
        restaurant.name = data.get('name', restaurant.name)
        restaurant.description = data.get('description', restaurant.description)
        restaurant.address = data.get('address', restaurant.address)
        restaurant.city = data.get('city', restaurant.city)
        restaurant.is_active = data.get('isActive', restaurant.is_active)
        restaurant.opening_hours = data.get('openingHours', restaurant.opening_hours)
        # Save the updated gallery list sent from the frontend
        restaurant.gallery = data.get('gallery', restaurant.gallery)
        
        db.session.commit()
        try:
            cache.delete_memoized(get_restaurant_details, restaurant.id)
            cache.delete_memoized(get_featured_restaurants)
        except Exception:
            pass
        
        return jsonify({"message": "Restaurant profile updated successfully!"}), 200
    


@app.route('/api/restaurant/promotions', methods=['GET', 'POST'])
@auth_required('token')
@roles_required('owner')
def manage_restaurant_promotions():
    """ Fetches (GET) or creates (POST) coupons for the owner's restaurant. """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()

    if request.method == 'GET':
        coupons = Coupon.query.filter_by(restaurant_id=restaurant.id).all()
        coupons_data = [{
            'id': c.id,
            'code': c.code,
            'type': c.discount_type,
            'value': c.discount_value,
            'isActive': c.is_active
        } for c in coupons]
        return jsonify(coupons_data), 200

    if request.method == 'POST':
        data = request.get_json()
        # Check for duplicate coupon codes for the same restaurant
        existing_coupon = Coupon.query.filter_by(restaurant_id=restaurant.id, code=data['code']).first()
        if existing_coupon:
            return jsonify({"message": f"Coupon code '{data['code']}' already exists for your restaurant."}), 409

        new_coupon = Coupon(
            restaurant_id=restaurant.id,
            code=data['code'],
            discount_type=data['type'],
            discount_value=data['value'],
            is_active=data.get('isActive', True)
        )
        db.session.add(new_coupon)
        db.session.commit()
        try:
            cache.delete_memoized(get_restaurant_details, restaurant.id)
            cache.delete_memoized(get_featured_restaurants)
        except Exception:
            pass
        return jsonify({"message": "Coupon created successfully."}), 201

@app.route('/api/restaurant/promotions/<int:coupon_id>', methods=['PUT', 'DELETE'])
@auth_required('token')
@roles_required('owner')
def manage_specific_promotion(coupon_id):
    """ Updates (PUT) or deletes (DELETE) a specific coupon. """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    coupon = Coupon.query.get_or_404(coupon_id)
    
    # Security check to ensure the coupon belongs to the owner's restaurant
    if coupon.restaurant_id != restaurant.id:
        return jsonify({"message": "Unauthorized to modify this coupon."}), 403

    if request.method == 'PUT':
        data = request.get_json()
        coupon.code = data.get('code', coupon.code)
        coupon.discount_type = data.get('type', coupon.discount_type)
        coupon.discount_value = data.get('value', coupon.discount_value)
        coupon.is_active = data.get('isActive', coupon.is_active)
        db.session.commit()
        try:
            cache.delete_memoized(get_restaurant_details, restaurant.id)
            cache.delete_memoized(get_featured_restaurants)
        except Exception:
            pass
        return jsonify({"message": "Coupon updated successfully."}), 200

    if request.method == 'DELETE':
        db.session.delete(coupon)
        db.session.commit()
        try:
            cache.delete_memoized(get_restaurant_details, restaurant.id)
            cache.delete_memoized(get_featured_restaurants)
        except Exception:
            pass
        return jsonify({"message": "Coupon deleted successfully."}), 200

# --- NEW: ANALYTICS ENDPOINT ---

@app.route('/api/restaurant/analytics', methods=['GET'])
@auth_required('token')
@roles_required('owner')
def get_restaurant_analytics():
    """ Gathers and returns all key analytics data for the owner's restaurant. """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    
    # --- Aggregate Stats ---
    total_revenue = db.session.query(func.sum(Order.total_amount))\
        .filter(Order.restaurant_id == restaurant.id, Order.status == 'completed').scalar() or 0.0
        
    total_orders = db.session.query(func.count(Order.id))\
        .filter(Order.restaurant_id == restaurant.id, Order.status == 'completed').scalar() or 0
        
    avg_order_value = total_revenue / total_orders if total_orders > 0 else 0.0

    stats = {
        'totalRevenue': round(total_revenue, 2),
        'totalOrders': total_orders,
        'avgOrderValue': round(avg_order_value, 2)
    }
    
    # --- Daily Sales (Last 7 Days) ---
    seven_days_ago = date.today() - timedelta(days=6)
    daily_sales_query = db.session.query(
            func.cast(Order.created_at, Date).label('order_date'),
            func.sum(Order.total_amount).label('daily_revenue')
        ).filter(
            Order.restaurant_id == restaurant.id,
            Order.status == 'completed',
            func.cast(Order.created_at, Date) >= seven_days_ago
        ).group_by('order_date').all()
    
    # Create a dictionary for easy lookup
    sales_by_date = {res.order_date: res.daily_revenue for res in daily_sales_query}
    
    daily_sales_data = []
    for i in range(7):
        current_date = date.today() - timedelta(days=i)
        daily_sales_data.append({
            'day': current_date.strftime('%b %d'),
            'sales': round(sales_by_date.get(current_date, 0), 2)
        })
    daily_sales_data.reverse() # Order from oldest to newest

    # --- Most Popular Items ---
    popular_items_query = db.session.query(
            MenuItem.name,
            func.count(OrderItem.id).label('order_count')
        ).join(OrderItem, MenuItem.id == OrderItem.menu_item_id)\
        .filter(MenuItem.restaurant_id == restaurant.id)\
        .group_by(MenuItem.name)\
        .order_by(func.count(OrderItem.id).desc()).limit(5).all()

    popular_items_data = [{'name': name, 'orders': count} for name, count in popular_items_query]

    return jsonify({
        'stats': stats,
        'dailySales': daily_sales_data,
        'popularItems': popular_items_data
    }), 200
# In backend/routes.py






@app.route('/api/menu-items/regular', methods=['GET'])
def get_regular_menu():
    # Fetch a larger pool of items (e.g. 50) to ensure we can find 6 unique ones
    all_items = MenuItem.query.limit(50).all()
    
    unique_items = []
    seen_names = set()
    
    for item in all_items:
        if item.name not in seen_names:
            seen_names.add(item.name)
            unique_items.append({
                'id': item.id, 
                'name': item.name, 
                'price': item.price, 
                'restaurantId': item.restaurant_id, 
                'reviews': 0, 
                'image': item.image_url or f'https://placehold.co/600x400/E65100/FFF?text={item.name.replace(" ", "+")}'
            })
            if len(unique_items) == 6:
                break
                
    return jsonify(unique_items), 200

@app.route('/api/favorites', methods=['GET'])
@auth_required('token')
@roles_required('customer')
def get_customer_favorites():
    favorites = current_user.favorites
    return jsonify([{'id': r.id, 'name': r.name} for r in favorites]), 200

# --- ✅ MODIFIED: ADMIN RESTAURANT CREATION/UPDATE ---
@app.route('/api/admin/restaurants', methods=['POST'])
@auth_required('token')
@roles_required('admin')
def admin_create_restaurant():
    data = request.get_json()
    # ... (owner validation remains the same) ...
    owner = user_datastore.find_user(email=data.get('ownerEmail'))
    if not owner or 'owner' not in [role.name for role in owner.roles]:
        return jsonify({"message": "Owner not found."}), 404

    new_restaurant = Restaurant(
        name=data.get('name'),
        address=data.get('address'),
        city=data.get('city'),
        owner_id=owner.id,
        latitude=data.get('latitude'),   # <-- ADDED
        longitude=data.get('longitude'), # <-- ADDED
        is_verified=True
    )
    db.session.add(new_restaurant)
    db.session.commit()
    try:
        cache.delete_memoized(get_featured_restaurants)
        cache.delete_memoized(get_restaurant_details, new_restaurant.id)
    except Exception:
        pass
    return jsonify({"message": "Restaurant created successfully."}), 201

@app.route('/api/admin/restaurants/<int:id>', methods=['PUT'])
@auth_required('token')
@roles_required('admin')
def admin_update_restaurant(id):
    restaurant = Restaurant.query.get_or_404(id)
    data = request.get_json()
    # ... (owner update logic remains the same) ...

    restaurant.name = data.get('name', restaurant.name)
    restaurant.address = data.get('address', restaurant.address)
    restaurant.city = data.get('city', restaurant.city)
    restaurant.latitude = data.get('latitude', restaurant.latitude)   # <-- ADDED
    restaurant.longitude = data.get('longitude', restaurant.longitude) # <-- ADDED
    
    db.session.commit()
    try:
        cache.delete_memoized(get_featured_restaurants)
        cache.delete_memoized(get_restaurant_details, restaurant.id)
    except Exception:
        pass
    return jsonify({"message": "Restaurant updated successfully."}), 200


@app.route('/api/admin/restaurants/export')
@auth_required('token')
@roles_required('admin')
def export_restaurants():
    try:
        # Subquery to calculate performance metrics
        order_metrics = db.session.query(
            Order.restaurant_id,
            func.count(Order.id).label('total_orders'),
            func.sum(Order.total_amount).label('total_revenue')
        ).filter(Order.status == 'completed').group_by(Order.restaurant_id).subquery()

        # Main query to get all restaurant data
        restaurants = db.session.query(
            Restaurant,
            func.coalesce(order_metrics.c.total_orders, 0).label('total_orders'),
            func.coalesce(order_metrics.c.total_revenue, 0).label('total_revenue')
        ).outerjoin(order_metrics, Restaurant.id == order_metrics.c.restaurant_id).all()

        # Create an Excel file in memory
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Restaurants"

        # Define headers
        headers = ["ID", "Name", "Owner Email", "City", "Status", "Total Orders", "Total Revenue"]
        sheet.append(headers)

        # Populate rows
        for restaurant, total_orders, total_revenue in restaurants:
            status = "Pending"
            if restaurant.is_verified and restaurant.owner and restaurant.owner.active:
                status = "Verified"
            elif restaurant.owner and not restaurant.owner.active:
                status = "Blocked"

            row = [
                restaurant.id,
                restaurant.name,
                restaurant.owner.email if restaurant.owner else 'N/A',
                restaurant.city,
                status,
                total_orders,
                round(float(total_revenue), 2)
            ]
            sheet.append(row)

        # Save the workbook to an in-memory stream
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='restaurants_export.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error exporting restaurants: {e}")
        return jsonify({"message": "Failed to export data."}), 500
    
@app.route('/api/admin/users/export')
@auth_required('token')
@roles_required('admin')
def export_users():
    try:
        # Re-using the subquery from get_all_users
        orders_subquery = db.session.query(
            Order.user_id,
            func.count(Order.id).label('total_orders'),
            func.sum(Order.total_amount).label('total_spent')
        ).group_by(Order.user_id).subquery()

        customer_role = Role.query.filter_by(name='customer').first()
        users = db.session.query(
            User,
            func.coalesce(orders_subquery.c.total_orders, 0).label('total_orders'),
            func.coalesce(orders_subquery.c.total_spent, 0).label('total_spent')
        ).outerjoin(orders_subquery, User.id == orders_subquery.c.user_id).filter(User.roles.contains(customer_role)).all()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Users"
        headers = ["ID", "Name", "Email", "Status", "Total Orders", "Total Spent"]
        sheet.append(headers)

        for user, total_orders, total_spent in users:
            status = "Active" if user.active else "Blocked"
            row = [
                user.id,
                user.name,
                user.email,
                status,
                total_orders,
                round(float(total_spent), 2)
            ]
            sheet.append(row)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='users_export.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error exporting users: {e}")
        return jsonify({"message": "Failed to export user data."}), 500
    
@app.route('/api/admin/orders/export')
@auth_required('token')
@roles_required('admin')
def export_orders():
    try:
        # Fetch all orders with related data
        orders = Order.query.options(
            joinedload(Order.customer),
            joinedload(Order.restaurant)
        ).order_by(Order.created_at.desc()).all()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Orders"

        headers = ["Order ID", "Customer Name", "Restaurant Name", "Date", "Total Amount", "Status", "Order Type"]
        sheet.append(headers)

        for order in orders:
            row = [
                order.id,
                order.customer.name if order.customer else 'N/A',
                order.restaurant.name if order.restaurant else 'N/A',
                order.created_at.strftime('%Y-%m-%d %H:%M'),
                order.total_amount,
                order.status.capitalize(),
                order.order_type.capitalize()
            ]
            sheet.append(row)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='orders_export.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error exporting orders: {e}")
        return jsonify({"message": "Failed to export order data."}), 500
@app.route('/api/admin/coupons/<int:coupon_id>/toggle', methods=['PATCH'])
@auth_required('token')
@roles_required('admin')
def admin_toggle_coupon_status(coupon_id):
    """ Toggles the activation status of a platform-wide coupon. """
    coupon = Coupon.query.filter_by(id=coupon_id, restaurant_id=None).first_or_404()
    
    # Flip the boolean status
    coupon.is_active = not coupon.is_active
    db.session.commit()
    
    new_status = "activated" if coupon.is_active else "deactivated"
    return jsonify({"message": f"Coupon '{coupon.code}' has been {new_status}."}), 200



# --- ✅ START: HAVERSINE FORMULA HELPER FUNCTION ---
def haversine(lat1, lon1, lat2, lon2):
    """
    Calculate the great-circle distance between two points 
    on the earth (specified in decimal degrees)
    """
    # convert decimal degrees to radians 
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])

    # haversine formula 
    dlon = lon2 - lon1 
    dlat = lat2 - lat1 
    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a)) 
    r = 6371 # Radius of earth in kilometers.
    return c * r
# --- ✅ END: HAVERSINE FORMULA HELPER FUNCTION ---

# --- ✅ START: NEW GEOLOCATION ENDPOINT ---
@app.route('/api/restaurants/nearby', methods=['GET'])
def get_nearby_restaurants():
    """
    Finds restaurants within a 7km radius of the user's location.
    Expects 'lat' and 'lng' as query parameters.
    """
    user_lat = request.args.get('lat', type=float)
    user_lng = request.args.get('lng', type=float)

    if user_lat is None or user_lng is None:
        return jsonify({"message": "Latitude and longitude are required."}), 400

    # Fetch all verified and active restaurants
    all_restaurants = Restaurant.query.filter_by(is_verified=True, is_active=True).all()
    
    nearby_restaurants = []
    for resto in all_restaurants:
        # Ensure the restaurant has location data
        if resto.latitude and resto.longitude:
            distance = haversine(user_lat, user_lng, resto.latitude, resto.longitude)
            if distance <= 7: # Check if within 7km radius
                nearby_restaurants.append(resto)

    # Format the data for the frontend
    restaurants_data = []
    for resto in nearby_restaurants:
        avg_rating = db.session.query(func.avg(Review.rating)).filter(Review.restaurant_id == resto.id).scalar() or 0.0
        review_count = db.session.query(func.count(Review.id)).filter(Review.restaurant_id == resto.id).scalar() or 0
        restaurants_data.append({
            'id': resto.id,
            'name': resto.name,
            'cuisine': 'Local Favorites', # You can enhance this later
            'rating': round(float(avg_rating), 1),
            'reviews': review_count,
            'image': f'https://placehold.co/600x400/E65100/FFF?text={resto.name.replace(" ", "+")}'
        })

    return jsonify(restaurants_data), 200
# --- ✅ END: NEW GEOLOCATION ENDPOINT ---



# ✅ THE FIX IS HERE: The @auth_required('token') decorator has been removed.
@app.route('/api/geocode', methods=['POST'])
def geocode_address():
    """
    Converts a physical address string into latitude and longitude coordinates
    using the free Nominatim (OpenStreetMap) API.
    """
    data = request.get_json()
    address = data.get('address')
    if not address:
        return jsonify({"message": "Address is required."}), 400

    # A custom User-Agent is required by Nominatim's terms of service.
    # Replace 'your-email@example.com' with your actual email if possible.
    headers = {'User-Agent': 'FoodleApp/1.0 (manimanjunath.v@gmail.com)'}
    params = {'q': address, 'format': 'json', 'limit': 1}
    
    try:
        response = requests.get('https://nominatim.openstreetmap.org/search', params=params, headers=headers)
        response.raise_for_status() # Raise an exception for bad status codes (like 4xx or 5xx)
        results = response.json()

        if not results:
            return jsonify({"message": "Address not found. Please try a different format or enter coordinates manually."}), 404

        location = results[0]
        return jsonify({
            "latitude": float(location['lat']),
            "longitude": float(location['lon'])
        }), 200

    except requests.exceptions.RequestException as e:
        print(f"Geocoding API error: {e}")
        return jsonify({"message": "Could not connect to the geocoding service. Please try again later or enter coordinates manually."}), 503



# --- ============================ ---
# --- NEW: CATEGORY CRUD ENDPOINTS ---
# --- ============================ ---

@app.route('/api/restaurant/categories', methods=['POST'])
@auth_required('token')
@roles_required('owner')
def create_category():
    """ Creates a new menu category for the owner's restaurant. """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({"message": "Category name is required."}), 400

    new_category = Category(
        name=data['name'],
        restaurant_id=restaurant.id
    )
    db.session.add(new_category)
    db.session.commit()
    return jsonify({"message": f"Category '{new_category.name}' created successfully."}), 201

@app.route('/api/restaurant/categories/<int:category_id>', methods=['PUT', 'DELETE'])
@auth_required('token')
@roles_required('owner')
def manage_category(category_id):
    """ Updates (PUT) or deletes (DELETE) a specific menu category. """
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    category = Category.query.get_or_404(category_id)

    # Security check: ensure the category belongs to the owner's restaurant
    if category.restaurant_id != restaurant.id:
        return jsonify({"message": "Unauthorized to modify this category."}), 403
        
    if request.method == 'PUT':
        data = request.get_json()
        if not data or not data.get('name'):
            return jsonify({"message": "Category name is required."}), 400
        category.name = data['name']
        db.session.commit()
        return jsonify({"message": "Category updated successfully."}), 200

    if request.method == 'DELETE':
        # Ensure category is empty before deleting
        if len(category.menu_items) > 0:
            return jsonify({"message": "Cannot delete a category that contains menu items. Please remove the items first."}), 409
        
        db.session.delete(category)
        db.session.commit()
        return jsonify({"message": "Category deleted successfully."}), 200


@app.route('/api/restaurant/menu/bulk-upload', methods=['POST'])
@auth_required('token')
@roles_required('owner')
def bulk_upload_menu():
    """
    Processes an Excel file to bulk-add categories and menu items.
    Expected Excel format: | Category | Name | Description | Price | Food Type (Veg/Non-Veg) |
    """
    if 'menu_file' not in request.files:
        return jsonify({"message": "No file part in the request."}), 400
    
    file = request.files['menu_file']

    if file.filename == '':
        return jsonify({"message": "No file selected."}), 400
        
    if not file.filename.endswith('.xlsx'):
        return jsonify({"message": "Invalid file type. Please upload a .xlsx file."}), 400

    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        category_cache = {} # Cache to avoid repeated DB lookups
        items_to_add = []
        items_added_count = 0

        # Iterate through rows, skipping the header
        for row in sheet.iter_rows(min_row=2, values_only=True):
            category_name, item_name, description, price, food_type = (list(row) + [None])[:5]
            if not category_name or not item_name or price is None:
                continue # Skip incomplete rows
            category_id = None
            if category_name in category_cache:
                category_id = category_cache[category_name]
            else:
                category = Category.query.filter_by(name=category_name, restaurant_id=restaurant.id).first()
                if category:
                    category_id = category.id
                else:
                    new_category = Category(name=category_name, restaurant_id=restaurant.id)
                    db.session.add(new_category)
                    db.session.flush() # Get the new ID before committing
                    category_id = new_category.id
                category_cache[category_name] = category_id

            new_item = MenuItem(
                name=item_name,
                description=description or "",
                price=float(price),
                category_id=category_id,
                restaurant_id=restaurant.id,
                food_type=food_type if food_type and food_type.strip().title() in ['Veg', 'Non-Veg'] else None
            )
            items_to_add.append(new_item)
            items_added_count += 1
        
        if items_to_add:
            db.session.add_all(items_to_add)
            db.session.commit()
            
        return jsonify({"message": f"Successfully added {items_added_count} menu items."}), 201

    except Exception as e:
        db.session.rollback()
        print(f"Error during bulk upload: {e}")
        return jsonify({"message": "An error occurred while processing the file. Please check the format and data."}), 500



@app.route('/api/restaurant/timeslots', methods=['GET', 'POST'])
@auth_required('token')
@roles_required('owner')
def manage_time_slots():
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    if request.method == 'GET':
        slots = TimeSlot.query.filter_by(restaurant_id=restaurant.id).order_by(TimeSlot.start_time).all()
        slots_data = [{'id': s.id, 'day_of_week': s.day_of_week, 'start_time': s.start_time.strftime('%H:%M'), 'end_time': s.end_time.strftime('%H:%M')} for s in slots]
        return jsonify(slots_data), 200
    if request.method == 'POST':
        data = request.get_json()
        new_slot = TimeSlot(
            restaurant_id=restaurant.id,
            day_of_week=data['day_of_week'],
            start_time=datetime.strptime(data['start_time'], '%H:%M').time(),
            end_time=datetime.strptime(data['end_time'], '%H:%M').time()
        )
        db.session.add(new_slot)
        db.session.commit()
        return jsonify({"message": "Time slot added successfully."}), 201

@app.route('/api/restaurant/timeslots/<int:slot_id>', methods=['DELETE'])
@auth_required('token')
@roles_required('owner')
def delete_time_slot(slot_id):
    restaurant = Restaurant.query.filter_by(owner_id=current_user.id).first_or_404()
    slot = TimeSlot.query.get_or_404(slot_id)
    if slot.restaurant_id != restaurant.id:
        return jsonify({"message": "Unauthorized"}), 403
    db.session.delete(slot)
    db.session.commit()
    return jsonify({"message": "Time slot deleted successfully."}), 200

@app.route('/api/restaurants/<int:restaurant_id>/available-slots', methods=['GET'])
@auth_required('token')
def get_available_slots(restaurant_id):
    """
    Calculates and returns a list of available time slots for the next 7 days
    for a given restaurant. This version is designed to be called once, providing all
    necessary data to the frontend.
    """
    try:
        # Fetch all defined time slots for the restaurant
        all_slots = db.session.query(TimeSlot).filter_by(restaurant_id=restaurant_id).all()
        
        # Group slots by day of the week for easy lookup
        slots_by_day = {}
        for slot in all_slots:
            if slot.day_of_week not in slots_by_day:
                slots_by_day[slot.day_of_week] = []
            slots_by_day[slot.day_of_week].append(slot)

        # Generate the available slots for the next 7 days
        available_days = []
        today = date.today()
        for i in range(7):
            current_date = today + timedelta(days=i)
            day_name = current_date.strftime('%A') # e.g., "Monday"

            if day_name in slots_by_day:
                day_slots = []
                for slot in sorted(slots_by_day[day_name], key=lambda x: x.start_time):
                    # Generate time slots in 30-minute increments within the defined window
                    current_time_dt = datetime.combine(current_date, slot.start_time)
                    end_time_dt = datetime.combine(current_date, slot.end_time)
                    
                    while current_time_dt < end_time_dt:
                        day_slots.append({
                            "value": current_time_dt.strftime('%Y-%m-%d %H:%M:%S'),
                            "display": current_time_dt.strftime('%I:%M %p').lstrip('0')
                        })
                        current_time_dt += timedelta(minutes=30)
                
                if day_slots:
                    available_days.append({
                        "date_value": current_date.strftime('%Y-%m-%d'),
                        "date_display": current_date.strftime('%A, %b %d'), # e.g., "Tuesday, Oct 14"
                        "slots": day_slots
                    })
                    
        return jsonify(available_days), 200

    except Exception as e:
        print(f"Error fetching available slots: {e}")
        return jsonify({"message": "An error occurred while fetching time slots."}), 500



@app.route('/api/orders/<int:order_id>/review', methods=['POST'])
@auth_required('token')
@roles_required('customer')
def submit_review(order_id):
    order = Order.query.filter_by(id=order_id, user_id=current_user.id).first_or_404()

    # --- Security & Business Logic Checks ---
    if order.status != 'completed':
        return jsonify({"message": "You can only review completed orders."}), 403
    
    if order.review:
        return jsonify({"message": "You have already submitted a review for this order."}), 409

    data = request.get_json()
    if not data or not data.get('rating'):
        return jsonify({"message": "Rating is a required field."}), 400

    new_review = Review(
        user_id=current_user.id,
        restaurant_id=order.restaurant_id,
        order_id=order.id,
        rating=data['rating'],
        comment=data.get('comment', '')
    )
    db.session.add(new_review)
    db.session.commit()

    return jsonify({"message": "Thank you for your review!"}), 201

@app.route('/api/restaurants/<int:restaurant_id>/reviews', methods=['GET'])
def get_restaurant_reviews(restaurant_id):
    reviews = Review.query.options(joinedload(Review.customer)).filter_by(restaurant_id=restaurant_id).order_by(Review.created_at.desc()).all()
    reviews_data = [{
        'id': r.id,
        'customerName': r.customer.name if r.customer else 'Anonymous',
        'rating': r.rating,
        'comment': r.comment,
        'date': r.created_at.strftime('%b %d, %Y')
    } for r in reviews]
    return jsonify(reviews_data), 200



@app.route('/api/upload/image', methods=['POST'])
@auth_required('token')
@roles_required('owner')
def upload_image():
    if 'image_file' not in request.files:
        return jsonify({"message": "No image file provided."}), 400
    
    file = request.files['image_file']

    if file.filename == '':
        return jsonify({"message": "No file selected."}), 400

    # Allow more input formats
    ext = file.filename.split('.')[-1].lower()
    if ext not in ['jpg', 'jpeg', 'png', 'webp']:
        return jsonify({"message": "Invalid file type. Please upload JPG, PNG, or WebP."}), 400
        
    # --- THIS IS THE KEY FIX ---
    # We create a new filename and ALWAYS give it a .jpg extension because we are converting to JPEG.
    unique_filename = f"{uuid.uuid4()}.jpg" 
    
    save_path = os.path.join(app.static_folder, 'assets', 'uploads', unique_filename)

    try:
        image = Image.open(file.stream)

        # Convert to RGB to remove transparency and ensure JPEG compatibility
        if image.mode in ("RGBA", "P"):
            image = image.convert("RGB")

        # Resize to a max width/height of 800px
        image.thumbnail((800, 800))
        
        # Save the compressed image as a JPEG
        image.save(save_path, 'jpeg', quality=85, optimize=True)

        # The public URL will now correctly point to the .jpg file
        public_url = f"/assets/uploads/{unique_filename}"
        
        return jsonify({"url": public_url}), 201

    except Exception as e:
        print(f"Error during image compression: {e}")
        return jsonify({"message": "An error occurred while processing the image."}), 500
@app.before_request
def debug_token():
    print("Received Authentication-Token:", request.headers.get('Authentication-Token'))

    

    
# --- ===================== ---
# --- CUSTOMER API RESOURCES ---
# --- ===================== ---
#api.add_resource(RestaurantListAPI, '/api/restaurants')
#api.add_resource(OrderAPI, '/api/orders')


# TODO: Add customer routes for reviews, favorites, rewards, etc.

# --- Frontend Serving Route ---
#@app.route('/', defaults={'path': ''})
#@app.route('/<path:path>')
#def serve_vue_app(path):

 #   return render_template('index.html')









